import sqlite3
from datetime import datetime
import calendar  
import math  # 📐 Used to compute perfect diagonal coordinate rotations for the watermark
from flask import Flask, render_template, request, redirect, make_response, session
from fpdf import FPDF
import io
from flask import send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = "super_secret_attendance_key_123"

# Custom template filter to easily format YYYY-MM-DD database strings to DD/MM/YYYY inside HTML templates
@app.template_filter('format_date')
def format_date_filter(value):
    if not value:
        return ""
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d/%m/%Y")
    except:
        return value

def cleanup_old_trash():
    try:
        conn = sqlite3.connect("database.db")
        cursor = conn.cursor()
        cursor.execute("""
            DELETE FROM employees 
            WHERE status = 'Deleted' 
            AND (julianday('now') - julianday(deleted_date)) >= 10
        """)
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error during auto-cleanup: {e}")


@app.route("/")
def dashboard():
    if not session.get("logged_in"):
        return redirect("/login")

    cleanup_old_trash()
    
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    
    today = datetime.now().strftime("%Y-%m-%d")
    current_month = datetime.now().strftime("%Y-%m")

    cursor.execute("SELECT COUNT(*) FROM employees WHERE status = 'Active'")
    total_employees = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM attendance WHERE attendance_date = ?", (today,))
    present_today = cursor.fetchone()[0]

    cursor.execute("""
        SELECT SUM(amount) FROM advances 
        WHERE strftime('%Y-%m', date) = ?
    """, (current_month,))
    advances_given = cursor.fetchone()[0]
    if advances_given is None:
        advances_given = 0.0

    cursor.execute("SELECT SUM(monthly_salary) FROM employees WHERE status = 'Active'")
    total_base_salary = cursor.fetchone()[0]
    if total_base_salary is None:
        total_base_salary = 0.0
        
    salary_due = max(0.0, float(total_base_salary) - float(advances_given))
    conn.close()

    return render_template(
        "dashboard.html",
        total_employees=total_employees,
        present_today=present_today,
        advances_given=advances_given,
        salary_due=salary_due
    )


@app.route("/employees")
def employees():
    if not session.get("logged_in"):
        return redirect("/login")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM employees WHERE status = 'Active'")
    employees = cursor.fetchall()
    conn.close()

    return render_template("employees.html", employees=employees)


@app.route("/add_employee", methods=["GET", "POST"])
def add_employee():
    if not session.get("logged_in"):
        return redirect("/login")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    if request.method == "POST":
        name = request.form["name"]
        
        raw_mobile = request.form["mobile"]
        mobile = "".join(filter(str.isdigit, raw_mobile))[:10]
        
        raw_aadhaar = request.form["aadhaar"]
        aadhaar = "".join(filter(str.isdigit, raw_aadhaar))[:12]

        joining_date = request.form["joining_date"]
        monthly_salary = request.form["monthly_salary"]
        shift_start = request.form["shift_start"]
        shift_end = request.form["shift_end"]
        grace_minutes = request.form["grace_minutes"]
        rfid_uid = request.form["rfid_uid"]

        cursor.execute("""
            INSERT INTO employees
            (
                name, mobile, aadhaar, joining_date, monthly_salary,
                shift_start, shift_end, grace_minutes, rfid_uid, status
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Active')
        """, (
            name, mobile, aadhaar, joining_date, monthly_salary,
            shift_start, shift_end, grace_minutes, rfid_uid
        ))
        
        cursor.execute("DELETE FROM pending_scan WHERE rfid_uid = ?", (rfid_uid,))
        conn.commit()
        conn.close()

        return redirect("/employees")

    if request.args.get("get_latest_scan") == "1":
        cursor.execute("SELECT rfid_uid FROM pending_scan WHERE id = 1")
        row = cursor.fetchone()
        conn.close()
        return {"rfid_uid": row[0] if row else ""}

    conn.close()
    return render_template("add_employee.html")


@app.route("/edit_employee/<int:id>", methods=["GET", "POST"])
def edit_employee(id):
    if not session.get("logged_in"):
        return redirect("/login")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    if request.method == "POST":
        name = request.form["name"]
        
        raw_mobile = request.form["mobile"]
        mobile = "".join(filter(str.isdigit, raw_mobile))[:10]
        
        raw_aadhaar = request.form["aadhaar"]
        aadhaar = "".join(filter(str.isdigit, raw_aadhaar))[:12]

        joining_date = request.form["joining_date"]
        monthly_salary = request.form["monthly_salary"]
        shift_start = request.form["shift_start"]
        shift_end = request.form["shift_end"]
        grace_minutes = request.form["grace_minutes"]
        rfid_uid = request.form["rfid_uid"]

        cursor.execute("""
            UPDATE employees
            SET
                name=?, mobile=?, aadhaar=?, joining_date=?, monthly_salary=?,
                shift_start=?, shift_end=?, grace_minutes=?, rfid_uid=?
            WHERE id=?
        """, (
            name, mobile, aadhaar, joining_date, monthly_salary,
            shift_start, shift_end, grace_minutes, rfid_uid, id
        ))
        conn.commit()
        conn.close()

        return redirect("/employees")

    cursor.execute("SELECT * FROM employees WHERE id=?", (id,))
    employee = cursor.fetchone()
    conn.close()

    return render_template("edit_employee.html", employee=employee)


@app.route("/delete_employee/<int:id>")
def delete_employee(id):
    if not session.get("logged_in"):
        return redirect("/login")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    today = datetime.now().strftime("%Y-%m-%d")

    cursor.execute("""
        UPDATE employees
        SET status = 'Deleted', deleted_date = ?
        WHERE id = ?
    """, (today, id))
    conn.commit()
    conn.close()

    return redirect("/employees")


@app.route("/trash")
def trash():
    if not session.get("logged_in"):
        return redirect("/login")

    cleanup_old_trash()

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM employees WHERE status = 'Deleted'")
    deleted_employees = cursor.fetchall()
    conn.close()

    return render_template("trash.html", employees=deleted_employees)


@app.route("/restore_employee/<int:id>")
def restore_employee(id):
    if not session.get("logged_in"):
        return redirect("/login")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE employees
        SET status = 'Active', deleted_date = NULL
        WHERE id = ?
    """, (id,))
    conn.commit()
    conn.close()

    return redirect("/trash")


@app.route("/attendance", methods=["GET", "POST"])
def attendance():
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    today = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H:%M:%S")

    if request.method == "POST":
        rfid_uid = str(request.form["rfid_uid"]).strip()

        cursor.execute("SELECT * FROM employees WHERE rfid_uid = ? AND status='Active'", (rfid_uid,))
        employee = cursor.fetchone()

        if employee:
            employee_id = employee[0]
            employee_name = employee[1]

            cursor.execute("""
                SELECT * FROM attendance WHERE employee_id=? AND attendance_date=?
            """, (employee_id, today))
            record = cursor.fetchone()

            if record is None:
                shift_start = employee[6]
                grace_minutes = int(employee[8])

                shift_time = datetime.strptime(shift_start, "%H:%M")
                check_time = datetime.strptime(current_time, "%H:%M:%S")

                allowed_minutes = (shift_time.hour * 60 + shift_time.minute + grace_minutes)
                actual_minutes = (check_time.hour * 60 + check_time.minute)
                late_minutes = max(0, actual_minutes - allowed_minutes)

                status = "On Time"
                if late_minutes > 0:
                    status = "Late"

                cursor.execute("""
                    INSERT INTO attendance (employee_id, attendance_date, check_in, status, late_minutes)
                    VALUES (?, ?, ?, ?, ?)
                """, (employee_id, today, current_time, status, late_minutes))
                conn.commit()
                conn.close()
                return f"{status}:{employee_name}"

            elif record[4] is None:
                check_in_time = datetime.strptime(record[3], "%H:%M:%S")
                check_out_time = datetime.strptime(current_time, "%H:%M:%S")

                difference = check_out_time - check_in_time
                hours = difference.seconds // 3600
                minutes = (difference.seconds % 3600) // 60
                working_hours = f"{hours}h {minutes}m"

                cursor.execute("""
                    UPDATE attendance SET check_out=?, working_hours=? WHERE id=?
                """, (current_time, working_hours, record[0]))
                conn.commit()
                conn.close()
                return f"CheckOut:{employee_name}"
                
            else:
                conn.close()
                return "ERROR:Duplicate Entry"
        else:
            cursor.execute("INSERT OR REPLACE INTO pending_scan (id, rfid_uid, timestamp) VALUES (1, ?, ?)", (rfid_uid, current_time))
            conn.commit()
            conn.close()
            return f"UNKNOWN:{rfid_uid}"

    cursor.execute("""
        SELECT attendance.id, employees.name, attendance.attendance_date, attendance.check_in,
               attendance.check_out, attendance.status, attendance.working_hours, attendance.late_minutes
        FROM attendance
        JOIN employees ON employees.id = attendance.employee_id
        ORDER BY attendance.id DESC
    """)
    attendance_records = cursor.fetchall()

    cursor.execute("SELECT id, name FROM employees WHERE status = 'Active'")
    active_employees = cursor.fetchall()
    
    conn.close()
    return render_template(
        "attendance.html", 
        attendance_records=attendance_records, 
        active_employees=active_employees
    )


@app.route("/manual_attendance", methods=["POST"])
def manual_attendance():
    if not session.get("logged_in"):
        return redirect("/login")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    employee_id = request.form["employee_id"]
    attendance_date = request.form["attendance_date"]
    action_type = request.form["action_type"]  
    manual_time = request.form["manual_time"]  

    if len(manual_time) == 5:
        manual_time += ":00"

    cursor.execute("SELECT name, shift_start, grace_minutes FROM employees WHERE id = ?", (employee_id,))
    employee = cursor.fetchone()
    
    if not employee:
        conn.close()
        return "Employee profile not found", 404

    employee_name, shift_start, grace_minutes = employee[0], employee[1], employee[2]

    cursor.execute("""
        SELECT * FROM attendance WHERE employee_id=? AND attendance_date=?
    """, (employee_id, attendance_date))
    record = cursor.fetchone()

    if action_type == "Check-In":
        if record is not None:
            conn.close()
            return f"<script>alert('Error: A check-in record already exists for {employee_name} on {attendance_date}!'); window.history.back();</script>"

        shift_time = datetime.strptime(shift_start, "%H:%M")
        check_time = datetime.strptime(manual_time, "%H:%M:%S")

        allowed_minutes = (shift_time.hour * 60 + shift_time.minute + int(grace_minutes))
        actual_minutes = (check_time.hour * 60 + check_time.minute)
        late_minutes = max(0, actual_minutes - allowed_minutes)

        status = "On Time" if late_minutes == 0 else "Late"

        cursor.execute("""
            INSERT INTO attendance (employee_id, attendance_date, check_in, status, late_minutes)
            VALUES (?, ?, ?, ?, ?)
        """, (employee_id, attendance_date, manual_time, status, late_minutes))
        
    elif action_type == "Check-Out":
        if record is None:
            conn.close()
            return f"<script>alert('Error: Cannot Check-Out. No Check-In record was found for {employee_name} on {attendance_date}. Mark Check-In first!'); window.history.back();</script>"
        if record[4] is not None:
            conn.close()
            return f"<script>alert('Error: {employee_name} has already checked out for this date!'); window.history.back();</script>"

        check_in_time = datetime.strptime(record[3], "%H:%M:%S")
        check_out_time = datetime.strptime(manual_time, "%H:%M:%S")
        
        if check_out_time < check_in_time:
            conn.close()
            return f"<script>alert('Error: Check-Out time cannot be earlier than Check-In time ({record[3]})!'); window.history.back();</script>"

        difference = check_out_time - check_in_time
        hours = difference.seconds // 3600
        minutes = (difference.seconds % 3600) // 60
        working_hours = f"{hours}h {minutes}m"

        cursor.execute("""
            UPDATE attendance SET check_out=?, working_hours=? WHERE id=?
        """, (manual_time, working_hours, record[0]))

    conn.commit()
    conn.close()
    return redirect("/attendance")


@app.route("/attendance_history", methods=["GET", "POST"])
def attendance_history():
    if not session.get("logged_in"):
        return redirect("/login")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM employees WHERE status = 'Active'")
    employees = cursor.fetchall()

    query = """
        SELECT attendance.id, employees.name, attendance.attendance_date, attendance.check_in,
               attendance.check_out, attendance.working_hours, attendance.late_minutes
        FROM attendance
        JOIN employees ON employees.id = attendance.employee_id
        WHERE 1=1
    """
    params = []

    if request.method == "POST":
        employee_id = request.form["employee_id"]
        from_date = request.form["from_date"]
        to_date = request.form["to_date"]

        if employee_id:
            query += " AND attendance.employee_id=?"
            params.append(employee_id)
        if from_date:
            query += " AND attendance.attendance_date>=?"
            params.append(from_date)
        if to_date:
            query += " AND attendance.attendance_date<=?"
            params.append(to_date)

    query += " ORDER BY attendance.attendance_date DESC"
    cursor.execute(query, params)
    records = cursor.fetchall()
    conn.close()

    return render_template("attendance_history.html", records=records, employees=employees)


@app.route("/advances", methods=["GET", "POST"])
def advances():
    if not session.get("logged_in"):
        return redirect("/login")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    if request.method == "POST":
        employee_id = request.form["employee_id"]
        amount = float(request.form["amount"])
        date_given = request.form["date"]
        reason = request.form["reason"]

        cursor.execute("""
            INSERT INTO advances (employee_id, amount, date, reason) VALUES (?, ?, ?, ?)
        """, (employee_id, amount, date_given, reason))
        conn.commit()
        conn.close()
        return redirect("/advances")

    cursor.execute("SELECT id, name FROM employees WHERE status = 'Active'")
    active_employees = cursor.fetchall()

    cursor.execute("""
        SELECT advances.id, employees.name, advances.amount, advances.date, advances.reason
        FROM advances
        JOIN employees ON employees.id = advances.employee_id
        ORDER BY advances.date DESC
    """)
    advance_records = cursor.fetchall()
    conn.close()

    today_str = datetime.now().strftime("%Y-%m-%d")

    return render_template("advances.html", employees=active_employees, advances=advance_records, current_date=today_str)


@app.route("/attendance_summary", methods=["GET", "POST"])
def attendance_summary():
    if not session.get("logged_in"):
        return redirect("/login")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    selected_month = datetime.now().strftime("%Y-%m")
    if request.method == "POST":
        selected_month = request.form["month"]

    cursor.execute("SELECT id, name, monthly_salary FROM employees WHERE status = 'Active'")
    employees = cursor.fetchall()
    summary_records = []

    for emp in employees:
        emp_id = emp[0]
        emp_name = emp[1]
        base_salary = emp[2]

        cursor.execute("""
            SELECT COUNT(*) FROM attendance WHERE employee_id = ? AND strftime('%Y-%m', attendance_date) = ?
        """, (emp_id, selected_month))
        present_days = cursor.fetchone()[0]

        cursor.execute("""
            SELECT COUNT(*) FROM attendance 
            WHERE employee_id = ? AND strftime('%Y-%m', attendance_date) = ? AND status = 'Late'
        """, (emp_id, selected_month))
        late_days = cursor.fetchone()[0]

        cursor.execute("""
            SELECT SUM(late_minutes) FROM attendance WHERE employee_id = ? AND strftime('%Y-%m', attendance_date) = ?
        """, (emp_id, selected_month))
        total_late_minutes = cursor.fetchone()[0] or 0

        summary_records.append({
            "id": emp_id,
            "name": emp_name,
            "base_salary": base_salary,
            "present_days": present_days,
            "late_days": late_days,
            "total_late_minutes": total_late_minutes
        })

    conn.close()

    return render_template("attendance_summary.html", summary_records=summary_records, selected_month=selected_month)


@app.route("/payroll", methods=["GET", "POST"])
def payroll():
    if not session.get("logged_in"):
        return redirect("/login")

    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()

    selected_month = datetime.now().strftime("%Y-%m")
    if request.method == "POST":
        selected_month = request.form["month"]

    # 🌟 Split month parameters to deduce total calendar days
    year_int, month_int = map(int, selected_month.split("-"))
    total_days_in_month = calendar.monthrange(year_int, month_int)[1]

    cursor.execute("SELECT id, name, monthly_salary FROM employees WHERE status = 'Active'")
    employees = cursor.fetchall()
    payroll_records = []

    for emp in employees:
        emp_id = emp[0]
        emp_name = emp[1]
        base_salary = float(emp[2])

        # 🌟 Calculate dynamic present days
        cursor.execute("""
            SELECT COUNT(*) FROM attendance 
            WHERE employee_id = ? AND strftime('%Y-%m', attendance_date) = ?
        """, (emp_id, selected_month))
        present_days = cursor.fetchone()[0]

        # 🌟 NEW PRORATION FORMULA: Base salary divided by calendar days, scaled by attendance
        earned_salary = (base_salary / total_days_in_month) * present_days

        cursor.execute("""
            SELECT SUM(amount) FROM advances WHERE employee_id = ? AND strftime('%Y-%m', date) = ?
        """, (emp_id, selected_month))
        total_advance = cursor.fetchone()[0] or 0.0
        
        # Deduct advance allocation from earned proration subtotal
        final_salary = max(0.0, earned_salary - total_advance)

        payroll_records.append({
            "id": emp_id,
            "name": emp_name,
            "base_salary": base_salary,
            "present_days": present_days,
            "earned_salary": round(earned_salary, 2),
            "total_advance": total_advance,
            "final_salary": round(final_salary, 2)
        })

    conn.close()
    return render_template("payroll.html", payroll_records=payroll_records, selected_month=selected_month, total_days=total_days_in_month)


@app.route("/export_excel")
def export_excel():
    if not session.get("logged_in"):
        return redirect("/login")

    wb = openpyxl.Workbook()
    
    ws1 = wb.active
    ws1.title = "Employee Directory"
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, mobile, aadhaar, monthly_salary, joining_date FROM employees WHERE status = 'Active'")
    employees = cursor.fetchall()
    
    headers1 = ["Employee ID", "Full Name", "Mobile Contact", "Aadhaar Number", "Base Salary (INR)", "Joining Date"]
    ws1.append(headers1)
    for emp in employees:
        formatted_join = datetime.strptime(emp[5], "%Y-%m-%d").strftime("%d/%m/%Y") if emp[5] else ""
        ws1.append([f"EMP-00{emp[0]}", emp[1], emp[2], emp[3], float(emp[4]), formatted_join])
        
    ws2 = wb.create_sheet(title="Attendance Logs")
    cursor.execute("""
        SELECT a.id, e.name, a.attendance_date, a.check_in, a.check_out, a.status, a.late_minutes 
        FROM attendance a 
        JOIN employees e ON a.employee_id = e.id
    """)
    attendance_rows = cursor.fetchall()
    headers2 = ["Log ID", "Employee Name", "Date", "Check In", "Check Out", "Shift Status", "Late Minutes"]
    ws2.append(headers2)
    for log in attendance_rows:
        formatted_log_date = datetime.strptime(log[2], "%Y-%m-%d").strftime("%d/%m/%Y") if log[2] else ""
        ws2.append([log[0], log[1], formatted_log_date, log[3], log[4] if log[4] else "Active Shift", log[5], log[6]])

    ws3 = wb.create_sheet(title="Cash Advances Ledger")
    cursor.execute("""
        SELECT ad.id, e.name, ad.amount, ad.date, ad.reason FROM advances ad 
        JOIN employees e ON ad.employee_id = e.id
    """)
    advance_rows = cursor.fetchall()
    headers3 = ["Transaction ID", "Employee Name", "Advance Amount (INR)", "Issue Date", "Reason / Notes"]
    ws3.append(headers3)
    for adv in advance_rows:
        formatted_adv_date = datetime.strptime(adv[3], "%Y-%m-%d").strftime("%d/%m/%Y") if adv[3] else ""
        ws3.append([adv[0], adv[1], float(adv[2]), formatted_adv_date, adv[4] if adv[4] else "—"])

    header_fill = PatternFill(start_color="182B49", end_color="182B49", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1')
    )

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        sheet.row_dimensions[1].height = 24
        
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 13)
            for cell in col:
                if cell.row != 1:
                    cell.border = thin_border
                    cell.font = Font(name="Segoe UI", size=10)
    conn.close()

    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    
    return send_file(
        excel_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="DEEP_ENTERPRISE_Master_Ledger.xlsx"
    )


@app.route("/download_payslip/<int:employee_id>/<string:month>")
def download_payslip(employee_id, month):
    if not session.get("logged_in"):
        return redirect("/login")

    try:
        conn = sqlite3.connect("database.db")
        cursor = conn.cursor()

        cursor.execute("""
            SELECT name, mobile, aadhaar, monthly_salary, joining_date, shift_start, shift_end 
            FROM employees WHERE id = ?
        """, (employee_id,))
        employee = cursor.fetchone()
        
        if not employee:
            conn.close()
            return "Employee not found", 404

        name, mobile, aadhaar = employee[0], employee[1], employee[2]
        base_salary = float(employee[3])
        joining_date, shift_start, shift_end = employee[4], employee[5], employee[6]

        cursor.execute("""
            SELECT COUNT(*) FROM attendance 
            WHERE employee_id = ? AND strftime('%Y-%m', attendance_date) = ?
        """, (employee_id, month))
        present_days = cursor.fetchone()[0]

        if present_days == 0:
            conn.close()
            return f"<script>alert('Error: Cannot generate payslip. {name} has 0 present days logged for {month}!'); window.history.back();</script>", 400

        year_int, month_int = map(int, month.split("-"))
        total_days_in_month = calendar.monthrange(year_int, month_int)[1]
        absent_days = max(0, total_days_in_month - present_days)

        # 🌟 NEW PRORATION PAY CALCULATOR: Compiles earned scale metrics
        earned_salary = (base_salary / total_days_in_month) * present_days

        cursor.execute("SELECT SUM(amount) FROM advances WHERE employee_id = ? AND strftime('%Y-%m', date) = ?", (employee_id, month))
        total_advance = cursor.fetchone()[0] or 0.0
        
        # Dynamic take home evaluation
        net_salary = max(0.0, earned_salary - float(total_advance))
        conn.close()

        # =========================================================================
        # 🎨 EXTENSION SUBCLASS FOR CROSSWISE ROTATED BACKGROUND WATERMARK LOOK
        # =========================================================================
        class WATERMARK_ENGINE_PDF(FPDF):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self.angle = 0

            def rotate(self, angle, x=-1, y=-1):
                if x == -1:
                    x = self.x
                if y == -1:
                    y = self.y
                if self.angle != 0:
                    self._out('Q')
                self.angle = angle
                if angle != 0:
                    rad = angle * math.pi / 180
                    c = math.cos(rad)
                    s = math.sin(rad)
                    cx = x * self.k
                    cy = (self.h - y) * self.k
                    self._out(f'q {c:.5f} {s:.5f} {-s:.5f} {c:.5f} {cx:.2f} {cy:.2f} cm 1 0 0 1 {-cx:.2f} {-cy:.2f} cm')

            def _endpage(self):
                if self.angle != 0:
                    self.angle = 0
                    self._out('Q')
                super()._endpage()

            def header(self):
                self.set_font('Arial', 'B', 46)
                self.set_text_color(240, 242, 246)  
                self.rotate(45, 105, 148)  
                self.text(32, 152, "DEEP ENTERPRISE")
                self.rotate(0)  

        pdf = WATERMARK_ENGINE_PDF()
        pdf.add_page()
        pdf.set_margins(15, 15, 15)
        
        pdf.set_fill_color(24, 43, 73)
        pdf.rect(15, 15, 180, 26, "F")
        
        try:
            pdf.image("static/logo.jpeg", x=18, y=17, w=22, h=22)
            pdf.set_xy(45, 15)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arial", style="B", size=16)
            pdf.cell(150, 16, txt="DEEP ENTERPRISE", border=0, align="L")
        except:
            pdf.set_xy(15, 15)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arial", style="B", size=16)
            pdf.cell(180, 16, txt="DEEP ENTERPRISE", border=0, align="C")

        pdf.set_xy(15, 27)
        pdf.set_font("Arial", style="I", size=9)
        pdf.cell(180, 10, txt="Automated Monthly Payroll Settlement Slip", border=0, align="C")
        pdf.ln(18)
        
        formatted_joining_date = datetime.strptime(joining_date, "%Y-%m-%d").strftime("%d/%m/%Y") if joining_date else ""
        formatted_gen_date = datetime.now().strftime('%d/%m/%Y')

        pdf.set_text_color(40, 40, 40)
        pdf.set_font("Arial", style="B", size=11)
        pdf.cell(90, 6, txt=f"Statement Payroll Cycle: {month}", border=0)
        pdf.cell(90, 6, txt=f"File Generation Date: {formatted_gen_date}", border=0, align="R")
        pdf.ln(10)
        
        pdf.set_draw_color(200, 200, 200)
        pdf.line(15, 54, 195, 54)
        pdf.ln(4)
        
        pdf.set_font("Arial", style="B", size=11)
        pdf.cell(180, 6, txt="EMPLOYEE DETAILS", border=0)
        pdf.ln(8)
        
        pdf.set_font("Arial", size=10)
        pdf.cell(35, 7, txt="Employee Name:", border=0)
        pdf.set_font("Arial", style="B", size=10)
        pdf.cell(55, 7, txt=str(name), border=0)
        pdf.set_font("Arial", style="B", size=10)
        pdf.cell(40, 7, txt="System Record ID:", border=0)
        pdf.cell(50, 7, txt=f"EMP-00{employee_id}", border=0)
        pdf.ln(7)
        
        pdf.set_font("Arial", size=10)
        pdf.cell(35, 7, txt="Mobile Contact:", border=0)
        pdf.cell(55, 7, txt=str(mobile), border=0)
        pdf.cell(40, 7, txt="Aadhaar Number:", border=0)
        pdf.cell(50, 7, txt=str(aadhaar), border=0)
        pdf.ln(7)
        
        pdf.cell(35, 7, txt="Assigned Shift:", border=0)
        pdf.cell(55, 7, txt=f"{shift_start} to {shift_end}", border=0)
        pdf.cell(40, 7, txt="Joining Date:", border=0)
        pdf.cell(50, 7, txt=str(formatted_joining_date), border=0)
        pdf.ln(10)
        
        pdf.set_draw_color(226, 232, 240)
        pdf.set_fill_color(248, 250, 252)
        pdf.set_font("Arial", style="B", size=10)
        pdf.cell(45, 8, txt="  Total Days in Month:", border=1, fill=True)
        pdf.set_font("Arial", size=10)
        pdf.cell(20, 8, txt=f" {total_days_in_month}", border=1, align="C")
        pdf.set_font("Arial", style="B", size=10)
        pdf.cell(40, 8, txt="  Days Present:", border=1, fill=True)
        pdf.set_font("Arial", size=10)
        pdf.cell(15, 8, txt=f" {present_days}", border=1, align="C")
        pdf.set_font("Arial", style="B", size=10)
        pdf.cell(45, 8, txt="  Days Absent:", border=1, fill=True)
        pdf.set_font("Arial", size=10)
        pdf.cell(15, 8, txt=f" {absent_days}", border=1, align="C")
        pdf.ln(16)
        
        pdf.set_text_color(40, 40, 40)
        pdf.set_font("Arial", style="B", size=11)
        pdf.cell(180, 6, txt="EARNINGS & DEDUCTIONS LEDGER STATEMENT", border=0)
        pdf.ln(8)
        
        pdf.set_fill_color(240, 244, 248)
        pdf.set_font("Arial", style="B", size=10)
        pdf.cell(120, 9, txt="  Transaction Item Description", border=1, fill=True)
        pdf.cell(60, 9, txt="Amount (INR)  ", border=1, align="R", fill=True)
        pdf.ln(9)
        
        pdf.set_font("Arial", size=10)
        pdf.cell(120, 9, txt="  Gross Basic Base Monthly Salary Allocation", border=1)
        pdf.cell(60, 9, txt=f"Rs. {base_salary:.2f}  ", border=1, align="R")
        pdf.ln(9)

        # 🌟 NEW ROW IN LEDGER BOX: Displays precise proration variables inside the itemized breakdown container
        pdf.set_font("Arial", size=10)
        pdf.cell(120, 9, txt=f"  Earned Prorated Salary Subtotal ({present_days}/{total_days_in_month} Days Present)", border=1)
        pdf.cell(60, 9, txt=f"Rs. {earned_salary:.2f}  ", border=1, align="R")
        pdf.ln(9)
        
        pdf.set_text_color(180, 40, 40)
        pdf.cell(120, 9, txt="  Less: Field Advance Deductions / Salary Advances Debited", border=1)
        pdf.cell(60, 9, txt=f"- Rs. {total_advance:.2f}  ", border=1, align="R")
        pdf.ln(9)
        
        pdf.set_text_color(24, 43, 73)
        pdf.set_font("Arial", style="B", size=11)
        pdf.set_fill_color(230, 240, 250)
        pdf.cell(120, 11, txt="  NET TAKE-HOME DISBURSEMENT AMOUNT", border=1, fill=True)
        pdf.cell(60, 11, txt=f"Rs. {net_salary:.2f}  ", border=1, align="R", fill=True)
        pdf.ln(22)
        
        pdf.set_text_color(60, 60, 60)
        pdf.set_font("Arial", style="I", size=10)
        pdf.cell(90, 10, txt="Authorized Manager Signature: __________________", border=0)
        pdf.cell(90, 10, txt="Receiver Employee Signature: __________________", border=0, align="R")
        pdf.ln(18)
        
        pdf.set_draw_color(230, 230, 230)
        pdf.line(15, 202, 195, 202)
        pdf.ln(4)
        pdf.set_font("Arial", size=8)
        pdf.set_text_color(130, 130, 130)
        pdf.cell(180, 5, txt="This document is an automated computer-generated processing summary statement.", border=0, align="C")
        pdf.ln(4)
        pdf.cell(180, 5, txt="DEEP ENTERPRISE Administration System logs verify the metrics compiled above directly.", border=0, align="C")

        response = make_response(bytes(pdf.output()))
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=Payslip_{name}_{month}.pdf'
        return response
    except Exception as e:
        return f"Error creating premium PDF payslip document layout: {str(e)}", 500


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username_input = request.form["username"]
        password_input = request.form["password"]
        if username_input == "admin" and password_input == "admin123":
            session["logged_in"] = True
            return redirect("/")
        else:
            error = "Invalid Username or Password. Please try again."
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.pop("logged_in", None)
    return redirect("/login")


if __name__ == "__main__":
    conn = sqlite3.connect("database.db")
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS advances (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER NOT NULL,
            amount REAL NOT NULL,
            date TEXT NOT NULL,
            reason TEXT,
            FOREIGN KEY (employee_id) REFERENCES employees (id)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS pending_scan (
            id INTEGER PRIMARY KEY,
            rfid_uid TEXT UNIQUE,
            timestamp TEXT
        )
    """)
    conn.commit()
    conn.close()

    app.run(host="0.0.0.0", port=5000, debug=True)